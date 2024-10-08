from copy import copy
from io import BytesIO

from django.http import HttpResponse
from openpyxl import load_workbook
from rest_framework.views import APIView

from students.models import Student


class ExcelData(APIView):

    def get(self, request, *args, **kwargs):
        # Path to the template file
        template_path = 'students/shablon.xlsx'

        # Load the workbook and select the first sheet ("ўқувчилар рўйхати")
        workbook = load_workbook(template_path)
        sheet = workbook["ўқувчилар рўйхати"]

        # Fetch all students from the database
        branch = request.query_params.get('branch')
        students = Student.objects.filter(user__branch__id=branch).all().order_by('user__surname', 'user__name')

        # Helper function to copy cell styles
        def copy_cell_style(source_cell, target_cell):
            if source_cell.has_style:  # Check if the source cell has a style
                target_cell.font = copy(source_cell.font)
                target_cell.border = copy(source_cell.border)
                target_cell.fill = copy(source_cell.fill)
                target_cell.number_format = copy(source_cell.number_format)
                target_cell.protection = copy(source_cell.protection)
                target_cell.alignment = copy(source_cell.alignment)

        # Row to use as a template for formatting (assuming row 2 is the template)
        template_row = 2

        # Starting row in the Excel sheet where the data will be written
        start_row = 3  # Start writing data from row 3 onwards (adjust based on your template)

        # Populate the Excel sheet with student data
        for idx, student in enumerate(students, start=start_row):
            row_num = idx

            # Copy styles from the template row to the current row
            for col_num in range(1, 16):  # Assuming you have 10 columns of data to fill
                source_cell = sheet.cell(row=template_row, column=col_num)
                target_cell = sheet.cell(row=row_num, column=col_num)
                copy_cell_style(source_cell, target_cell)

            # Add the data to the newly created row
            sheet[f"A{row_num}"] = idx - start_row + 1
            sheet[f"B{row_num}"] = student.region
            sheet[f"C{row_num}"] = student.district
            sheet[f"D{row_num}"] = student.old_school
            sheet[f"E{row_num}"] = f"{student.user.name} {student.user.surname}"
            sheet[f"F{row_num}"] = student.born_date
            sheet[f"G{row_num}"] = student.student_seria
            sheet[f"H{row_num}"] = student.student_seria_num
            sheet[f"I{row_num}"] = student.parents_fullname
            sheet[f"J{row_num}"] = f"{student.parent_region}"  # Яшаш манзили
            sheet[f"K{row_num}"] = f"{student.parent_seria}"  # Ota ona paspor
            sheet[f"L{row_num}"] = f"{student.parent_seria_num}"  # Яшаш манзили
            sheet[f"M{row_num}"] = f"{student.born_date}"  # Яшаш манзили
            sheet[f"N{row_num}"] = f"{student.parents_number}"

        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        response = HttpResponse(buffer,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=students.xlsx'

        return response
